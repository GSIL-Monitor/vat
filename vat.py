# -*- coding: utf-8 -*-
import sys
import os
import subprocess
import datetime
import shutil
import time
import logging
import codecs

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("Please use command to install openpyxl: pip install openpyxl")
    sys.exit(0)
try:
    from PIL import ImageGrab
except ModuleNotFoundError:
    print("Please use command to install pillow: pip install Pillow")
    sys.exit(0)

from configparser import ConfigParser
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QMainWindow, QWidget, QTextEdit
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtGui import QStandardItem
from PyQt5.QtGui import QTextCursor
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtGui import QTextOption
from PyQt5.QtCore import QUrl
from PyQt5.QtCore import Qt, QRect, QFileInfo
from PyQt5.QtCore import QThread
from PyQt5.QtCore import QObject
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QAction, QMenu
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import QCoreApplication
from vatgui import Ui_MainWindow
import textedit

import ctypes
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")  # 设置任务栏图标

__version__ = "1.0.0.2"
VAR_SEPARATOR = '\\'
VAR_EXCEL_REPORT_NAME = 'MTBF_Test_Report'
VAR_EXCEL_SUFFIX = '.xlsx'
VAR_EXCEL_NAME = VAR_EXCEL_REPORT_NAME + VAR_EXCEL_SUFFIX
VAR_SHEET_TIME = "MTBFTimePerSection"
VAR_SHEET_COMPLETION = "CompletionPercentage"
VAR_SHEET_DETAILSHEET = 'DetailSheet'
VAR_TIME_START_ROW = 10
VAR_TIME_START_COLUMN = 1
VAR_COMPLETION_START_ROW = 11
VAR_COMPLETION_START_COLUMN = 2
VAR_CURRENT_PATH = os.path.abspath(os.curdir) + VAR_SEPARATOR
VAR_FOLDER_RESOURCE = "resource"
VAR_FOLDER_REPORT = "report"
VAR_FOLDER_CASE = 'case'
VAR_FOLDER_CONFIGURE = 'config'
VAR_CONFIGURE_SUFFIX = '.ini'
VAR_LOG_FILE_SUFFIX = ".log"
VAR_DEFAULT_FILE_LOG = "log{0}.log".format(time.strftime("_%Y%m%d_%H%M%S"))
VAR_DEFAULT_IMAGE_SUFFIX = '.jpg'
VAR_CONFIG_ITEM_CASE = 'case'
VAR_CONFIG_ITEM_PYTHON = 'python'
VAR_CONFIG_SECTION_CONFIG = 'Config'
VAR_CONFIG_PATH = "{0}{1}{2}{3}{4}".format(VAR_CURRENT_PATH, VAR_FOLDER_CONFIGURE, VAR_SEPARATOR, VAR_FOLDER_CONFIGURE,
                                           VAR_CONFIGURE_SUFFIX)


# Unused current
def create_logger(name=__name__, file=VAR_DEFAULT_FILE_LOG):
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)
    file_handle = logging.FileHandler(file)
    file_handle.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s.%(msecs)03d: [%(name)s] [%(levelname)s] [%(funcName)s] %(message)s',
                                  '%y%m%d %H:%M:%S')
    file_handle.setFormatter(formatter)

    console_handle = logging.StreamHandler()
    console_handle.setLevel(logging.DEBUG)
    logger.addHandler(file_handle)
    logger.addHandler(console_handle)
    return logger


def save_image(name):
    im = ImageGrab.grab()
    im.save(name)


class GetConfig(object):
    def __init__(self):
        self.config = ConfigParser()
        # self.config_path = "{0}{1}{2}{3}{4}".format(VAR_CURRENT_PATH, VAR_FOLDER_CONFIGURE, VAR_SEPARATOR,
        #                                             VAR_FOLDER_CONFIGURE, VAR_CONFIGURE_SUFFIX)
        self.config.read(VAR_CONFIG_PATH)

    def get_int(self, section, key, default=0):
        try:
            return self.config.getint(section, key)
        except BaseException as e:
            return default

    def get_str(self, section, key, default=None):
        try:
            return self.config.get(section, key)
        except BaseException as e:
            return default

    def set_str(self, section, key, value):
        self.config.set(section, key, value)
        self.config.write(open(VAR_CONFIG_PATH, 'w'))

    def get_object(self):
        return self.config


class CaseCreator(object):
    def __init__(self, case_name):
        self.case_name = case_name
        self.time_start = None
        self.time_end = None
        self.loop_total = 0
        self.loop_current = 0
        self.result = None
        self.duration = None
        self.report = None
        self.fail_loop = None
        self.success_rate = 0
        self.count_fail = 0
        self.count_success = 0
        self.round_ = 0
        self.round_total = 0
        self.case_index = 0
        self.case_log = None


class Report(object):
    def __init__(self, excel_name=VAR_EXCEL_NAME):
        self.excel_name = excel_name
        self.wb = load_workbook(self.excel_name)
        self.sheet_time = self.wb[VAR_SHEET_TIME]
        self.sheet_completion = self.wb[VAR_SHEET_COMPLETION]
        self.sheet_detail = self.wb[VAR_SHEET_DETAILSHEET]

    def write_time(self, row, column, value):
        self.sheet_time.cell(row=row, column=column).value = value

    def write_completion(self, row, column, value):
        self.sheet_completion.cell(row=row, column=column).value = value

    def write_detail(self, row, column, value):
        self.sheet_detail.cell(row=row, column=column).value = value

    def save(self):
        self.wb.save(self.excel_name)


class RunThread(QThread):
    signal_test_finish = pyqtSignal()
    signal_case_info_update = pyqtSignal(CaseCreator)

    def __init__(self, parent=None):
        super(RunThread, self).__init__(parent)
        self.case_list = None
        self.stop_flag = False
        self.pipe = None
        self.report = None
        self.python_path = None
        # self.config = GetConfig

    def get_python_path(self):
        config = GetConfig()
        self.python_path = config.get_str(VAR_CONFIG_SECTION_CONFIG, VAR_CONFIG_ITEM_PYTHON)
        if not self.python_path:
            print("python path error, please click menu [setting]->[python] to set python path.")
            return False
        return True

    def run(self):
        self.detail_report_row = 3  # use for sheet detail
        self.stop_flag = False
        if self.case_list is None:
            return
        if not self.get_python_path():
            return

        self.report_folder_path = "{0}{1}{2}{3}{4}".format(VAR_CURRENT_PATH, VAR_FOLDER_REPORT, VAR_SEPARATOR,
                                                           VAR_EXCEL_REPORT_NAME, time.strftime("_%Y%m%d_%H%M%S"))
        self.create_folder(self.report_folder_path)
        self.create_folder(self.report_folder_path + "\\Log")

        report_source = "{0}{1}{2}{3}".format(VAR_CURRENT_PATH, VAR_FOLDER_RESOURCE, VAR_SEPARATOR, VAR_EXCEL_NAME)
        report_destination = "{0}{1}{2}{3}{4}".format(self.report_folder_path, VAR_SEPARATOR, VAR_EXCEL_REPORT_NAME,
                                                      time.strftime("_%Y%m%d_%H%M%S"), VAR_EXCEL_SUFFIX)
        self.copy_report(report_source, report_destination)
        self.report = Report(report_destination)
        config = GetConfig()
        test_round = config.get_int('Round', 'round', 6)
        for round_ in range(test_round):
            if self.stop_flag is True:
                break

            self.performance_case_list(round_)
            self.detail_report_row += 1
        self.test_finish()

    def test_finish(self):
        self.report.save()
        print("%s Test Finished %s" % ('-'*20, '-'*20))
        self.signal_test_finish.emit()

    @classmethod
    def copy_report(cls, source, destination):
        if os.path.exists(source):
            shutil.copyfile(source, destination)

    @classmethod
    def create_folder(cls, path):
        if not os.path.exists(path):
            os.makedirs(path)

    def performance_case_list(self, round_):
        config = GetConfig()
        for case in self.case_list:
            case_name = (case.split(VAR_SEPARATOR)[-1]).split('.')[0]
            case_loop = config.get_int('Loop', case_name)
            if not case_loop:
                case_loop = config.get_int("Loop", 'default_loop')
            self.performance_signal_case(case, case_loop, round_, self.case_list.index(case))
            self.detail_report_row += 1

    def performance_signal_case(self, case, loop, round_, case_index):
        case_creator = CaseCreator(case.split(VAR_SEPARATOR)[-1])
        case_creator.round_ = round_
        case_creator.case_index = case_index
        case_creator.loop_total = loop
        case_creator.time_start = datetime.datetime.now()
        case_creator.case_log_folder = "{0}{1}Log{2}Round_{3}_{4}".format(
            self.report_folder_path,
            VAR_SEPARATOR,
            VAR_SEPARATOR,
            str(case_creator.round_ + 1),
            str(case_creator.case_name).split('.')[0]
        )
        self.create_folder(case_creator.case_log_folder)
        case_creator.case_log = "{0}{1}Log{2}Round_{3}_{4}{5}Round_{6}_{7}".format(
            self.report_folder_path,
            VAR_SEPARATOR,
            VAR_SEPARATOR,
            str(case_creator.round_ + 1),
            str(case_creator.case_name).split('.')[0],
            VAR_SEPARATOR,
            str(case_creator.round_ + 1),
            str(case_creator.case_name).split('.')[0]
        )

        log_content = ""
        command = self.python_path + " " + case

        for current_loop in range(loop):
            if self.stop_flag is True:
                break
            case_creator.loop_current = current_loop
            self.signal_case_info_update.emit(case_creator)
            text = "{0} round: {1} count: {2} {3}".format('-'*20, case_creator.round_ + 1, current_loop + 1, '-'*20)
            self.write_log(case_creator.case_log + VAR_LOG_FILE_SUFFIX, text)
            print(text)

            self.pipe = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
            while self.pipe.poll() is None:
                line = self.pipe.stdout.readline()
                if line:
                    try:
                        log_content += (str(line, encoding='utf-8'))
                        print(str(line, encoding='utf-8'))
                    except:
                        log_content += (str(line, encoding='GB2312'))
                        print(str(line, encoding='GB2312'))

            if self.pipe.returncode == 0:
                case_creator.count_success += 1
                # log_content = ""  # clear log if test pass
            else:
                case_creator.count_fail += 1
                case_creator.fail_loop = current_loop
                save_image("{0}_FailLoop_{1}_{2}".format(case_creator.case_log, case_creator.fail_loop + 1,
                                                         VAR_DEFAULT_IMAGE_SUFFIX))

            case_creator.time_end = datetime.datetime.now()
            case_creator.success_rate = int((case_creator.count_success / loop) * 100)
            consumption = case_creator.time_end - case_creator.time_start
            case_creator.duration = self.remove_microsecond(consumption)

            self.write_report(case_creator)
            self.write_log(case_creator.case_log + VAR_LOG_FILE_SUFFIX, log_content)
            log_content = ""  # clear log if test finished
            self.signal_case_info_update.emit(case_creator)

    @classmethod
    def write_log(cls, file, text):
        if not text:
            return

        with open(file, 'a+') as f:
            f.write(text if text.endswith('\n') else text + '\n')

    def write_report(self, case_creator):
        sheet_time_row = VAR_TIME_START_ROW + case_creator.case_index
        sheet_completion_row = VAR_COMPLETION_START_ROW + case_creator.case_index
        # write value to sheet time
        self.report.write_time(sheet_time_row, VAR_TIME_START_COLUMN, case_creator.case_name)
        self.report.write_time(sheet_time_row, VAR_TIME_START_COLUMN + 1 + case_creator.round_, case_creator.duration)

        # write value to sheet completion
        self.report.write_completion(sheet_completion_row, VAR_COMPLETION_START_COLUMN, case_creator.case_name)
        self.report.write_completion(sheet_completion_row, VAR_COMPLETION_START_COLUMN + 1, case_creator.loop_total)
        self.report.write_completion(sheet_completion_row, VAR_COMPLETION_START_COLUMN + 2 + case_creator.round_,
                                     case_creator.count_success)

        # write value to sheet detail
        self.report.write_detail(row=self.detail_report_row, column=1, value=case_creator.round_ + 1)
        self.report.write_detail(row=self.detail_report_row, column=2, value=case_creator.case_name)
        self.report.write_detail(row=self.detail_report_row, column=3, value=case_creator.duration)
        self.report.write_detail(row=self.detail_report_row, column=4, value=case_creator.count_success)
        self.report.write_detail(row=self.detail_report_row, column=5, value=case_creator.count_fail)
        self.report.write_detail(row=self.detail_report_row, column=6, value=case_creator.loop_current)
        self.report.write_detail(row=self.detail_report_row, column=7,
                                 value='Passed' if case_creator.count_fail == 0 else "Failed")
        self.report.save()

    @classmethod
    def remove_microsecond(cls, text):
        text = str(text)
        position = text.find('.')
        return text[:position]

    def stop(self):
        self.stop_flag = True
        if self.pipe:
            self.pipe.terminate()

        if self.report:
            self.report.save()


class EmittingStream(QObject):
    text_written = pyqtSignal(str)

    def write(self, text):
        self.text_written.emit(str(text))


class VatWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        Ui_MainWindow.__init__(self)
        self.setupUi(self)
        self.run_case_list = list()
        self.thread = RunThread()
        self.table_model = QStandardItemModel(self.tableViewCaseRunDetail)
        self.model = QStandardItemModel()

        # initial ui
        self.init_ui()

        # initial treeview
        self.init_treeview()

        # initial table view
        self.init_case_status()

        # initial signal
        self.init_signal()

        sys.stdout = EmittingStream(text_written=self.output_written)
        sys.stderr = EmittingStream(text_written=self.output_written)

    def init_treeview(self):
        self.case_path = self.get_case_path()
        self.update_tree_case(self.case_path)
        # self.model.setHorizontalHeaderLabels(["Test Case"])
        # self.case_list.setModel(self.model)

    def init_signal(self):
        self.model.itemChanged.connect(self.check_tree_change)
        self.actionOpen.triggered.connect(self.action_open)
        self.actionPython.triggered.connect(self.action_python)
        self.actionRun.triggered.connect(self.action_run)
        self.actionStop.triggered.connect(self.action_stop)
        # self.actionAdd.triggered.connect(self.add_test_case_list)
        self.actionSetting.triggered.connect(self.action_setting)
        self.actionReport.triggered.connect(self.action_report)
        self.actionSave_2.triggered.connect(self.action_save)
        self.actionAbout.triggered.connect(self.action_about)
        self.actionExit.triggered.connect(self.action_exit)
        self.actionSynchronize.triggered.connect(self.action_synchronize)
        self.case_list.clicked.connect(self.case_tree_click)
        self.case_list.expanded.connect(self.case_tree_click)
        self.thread.signal_test_finish.connect(self.finish_test)
        self.thread.signal_case_info_update.connect(self.update_status)
        self.case_list.customContextMenuRequested.connect(self.show_context_menu)
        self.tabWidget.tabCloseRequested.connect(self.tab_close)

    def init_ui(self):
        self.actionStop.setEnabled(False)
        self.textEditLog.setReadOnly(True)
        self.textEditLog.setWordWrapMode(QTextOption.NoWrap)
        # remove config tab
        self.tabWidget.removeTab(1)
        self.tabWidget.setTabsClosable(True)
        # support right menu
        self.case_list.setContextMenuPolicy(Qt.CustomContextMenu)

    def show_context_menu(self, pos):
        self.right_menu = QMenu()
        self.right_open = QAction()
        self.right_open.setObjectName("right_open")
        self.right_open.setText("Open")
        self.right_menu.addAction(self.right_open)
        self.right_open.triggered.connect(self.right_action_open)

        self.right_menu.exec_(self.case_list.mapToGlobal(pos))

    def right_action_open(self):
        index = self.case_list.selectionModel().currentIndex()
        filename = self.case_path + self.get_all_parent_index(index)
        self.create_edit_tab(filename)

    def create_edit_tab(self, filename):
        if os.path.isdir(filename):
            return
        if filename:
            for i in range(self.tabWidget.count()):
                textEdit = self.tabWidget.widget(i)
                print(textEdit.windowTitle().title())
                print(filename)
                if textEdit.windowTitle().title().lower() == filename.split(VAR_SEPARATOR)[-1].lower():
                    self.tabWidget.setCurrentWidget(textEdit)
                    break
            else:
                self.load_file(filename)

    def load_file(self, filename):
        textEdit = textedit.TextEdit(filename)
        try:
            textEdit.load()
        except EnvironmentError as e:
            QMessageBox.warning(self,
                    "Tabbed Text Editor -- Load Error",
                    "Failed to load {0}: {1}".format(filename, e))
            textEdit.close()
            del textEdit
        else:
            self.tabWidget.addTab(textEdit, textEdit.windowTitle())
            self.tabWidget.setCurrentWidget(textEdit)
            textEdit.setWordWrapMode(QTextOption.NoWrap)

    def tab_close(self, index):
        if index == 0:
            return
        textEdit = self.tabWidget.currentWidget()
        if textEdit is None or not isinstance(textEdit, QTextEdit):
            return
        textEdit.close()

    def action_save(self):
        textEdit = self.tabWidget.currentWidget()
        if textEdit is None or not isinstance(textEdit, QTextEdit):
            return True
        try:
            textEdit.save()
            self.tabWidget.setTabText(self.tabWidget.currentIndex(),
                                      QFileInfo(textEdit.filename).fileName())
            return True
        except EnvironmentError as e:
            QMessageBox.warning(self,
                                "Tabbed Text Editor -- Save Error",
                                "Failed to save {0}: {1}".format(textEdit.filename, e))
            return False
        pass

    def action_synchronize(self):
        case_path = self.get_case_path()
        self.update_tree_case(case_path)
        pass

    def update_status(self, case_creator):
        status_loop = "{0}/{1}".format(case_creator.loop_current + 1, case_creator.loop_total)
        status_round = "{0}".format(case_creator.round_ + 1)
        self.table_model.setItem(case_creator.case_index, 1, QStandardItem(status_round))
        self.table_model.setItem(case_creator.case_index, 2, QStandardItem(status_loop))
        pass

    def action_exit(self):
        result = QMessageBox.question(self, "Exit", "Quit?", QMessageBox.Yes | QMessageBox.No)
        if result == QMessageBox.Yes:
            QCoreApplication.instance().quit()

    def action_about(self):
        text = "Version: {0}\nJRD COMMUNICATION Inc Copyright(c)2018\n".format(__version__)
        QMessageBox.about(self, "About", text)

    # def action_save(self):
    #     to_save_content = self.textEditConfig.toPlainText()
    #     # print(to_save_content)
    #     config = codecs.open(VAR_CONFIG_PATH, 'w', 'utf-8')
    #     config.write(to_save_content)
    #     pass

    def action_setting(self):
        self.create_edit_tab(VAR_CONFIG_PATH)
        # self.tabWidget.setCurrentIndex(1)
        # config_content = codecs.open(VAR_CONFIG_PATH, 'r', 'utf-8').read()
        # self.textEditConfig.setPlainText(config_content)

    def case_tree_click(self, index):
        self.case_list.resizeColumnToContents(0)
        pass

    @classmethod
    def action_report(cls):
        QDesktopServices.openUrl(QUrl.fromLocalFile(VAR_CURRENT_PATH + VAR_FOLDER_REPORT))

    def action_python(self):
        file, file_type = QFileDialog.getOpenFileName(self, "Select Python File", "C:\\", "Files(*.exe);;All Files(*)")
        print(file)
        if file:
            config = GetConfig()
            config.set_str(VAR_CONFIG_SECTION_CONFIG, VAR_CONFIG_ITEM_PYTHON, file)

    def get_case_path(self):
        config = GetConfig()
        path = config.get_str(VAR_CONFIG_SECTION_CONFIG, VAR_CONFIG_ITEM_CASE)
        if path:
            path = str(path).replace("/", VAR_SEPARATOR)
            if not path.endswith(VAR_SEPARATOR):
                path += VAR_SEPARATOR
        else:
            path = VAR_CURRENT_PATH + VAR_FOLDER_CASE + VAR_SEPARATOR
        return path

    def action_open(self):
        config = GetConfig()
        default_case_path = config.get_str(VAR_CONFIG_SECTION_CONFIG, VAR_CONFIG_ITEM_CASE)
        if not default_case_path:
            default_case_path = VAR_CURRENT_PATH

        path = QFileDialog.getExistingDirectory(self, "Select Case Folder", default_case_path)
        self.update_tree_case(path)
        # if path:
        #     path = str(path).replace("/", VAR_SEPARATOR)
        #     if not str(path).endswith(VAR_SEPARATOR):
        #         path += VAR_SEPARATOR
        #     self.case_path = path
        #     self.model.clear()
        #     self.model.setHorizontalHeaderLabels(["Test Case"])
        #     self.case_list.setModel(self.model)
        #     self.create_tree(self.model, self.case_path)
        #     config.set_str("Config", VAR_CONFIG_ITEM_CASE, path)
        #
        # self.clear_case_list()

    def update_tree_case(self, path):
        config = GetConfig()
        if path:
            path = str(path).replace("/", VAR_SEPARATOR)
            if not str(path).endswith(VAR_SEPARATOR):
                path += VAR_SEPARATOR
            self.case_path = path
            self.model.clear()
            self.model.setHorizontalHeaderLabels(["Test Case"])
            self.case_list.setModel(self.model)
            self.create_tree(self.model, self.case_path)
            config.set_str("Config", VAR_CONFIG_ITEM_CASE, path)

        self.clear_case_list()

    def clear_case_list(self):
        self.run_case_list.clear()

    def init_case_status(self):
        head_list = ['Name', 'Round', 'Loop', 'Time', 'Result', 'Success Rate', 'Status']
        self.table_model.setHorizontalHeaderLabels(head_list)
        self.tableViewCaseRunDetail.setModel(self.table_model)

    # def add_test_case_list(self):
    #     i = 0
    #     for case in self.run_case_list:
    #         case = case.split(VAR_SEPARATOR)[-1]
    #         self.add_case_to_table(i, case)
    #         i += 1
    #
    # def add_case_to_table(self, row, name):
    #     self.add_item_to_table(row, 0, name)
    #
    # def add_item_to_table(self, row, column, item):
    #     self.table_model.setItem(row, column, QStandardItem(item))
    #
    # def remove_case_from_table(self, case):
    #     row_count = self.table_model.rowCount()
    #     case = case.split(VAR_SEPARATOR)[-1]
    #     for count in range(row_count):
    #         case_name = self.table_model.data(self.table_model.index(count, 0))
    #         if case.__contains__(case_name):
    #             self.table_model.removeRow(count)
    #             break

    def action_run(self):
        self.tabWidget.setCurrentIndex(0)
        self.control_status()
        self.clear_log()
        self.thread.case_list = self.run_case_list
        self.thread.start()

    def action_stop(self):
        # self.control_status(run=False)
        self.actionStop.setEnabled(False)
        self.thread.stop()

    def finish_test(self):
        self.control_status(run=False)

    def control_status(self, run=True):
        self.actionRun.setEnabled(not run)
        self.actionStop.setEnabled(run)
        self.case_list.setEnabled(not run)
        self.actionOpen.setEnabled(not run)
        self.actionSave_2.setEnabled(not run)
        self.actionSetting.setEnabled(not run)
        self.actionPython.setEnabled(not run)
        self.actionReport.setEnabled(not run)
        self.actionExit.setEnabled(not run)
        self.actionSynchronize.setEnabled(not run)

    def output_written(self, text):
        cursor = self.textEditLog.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(text)
        self.textEditLog.setTextCursor(cursor)
        self.textEditLog.ensureCursorVisible()

    def clear_log(self):
        self.textEditLog.clear()

    def check_tree_change(self, item):
        if item.isCheckable():
            state = item.checkState()
            # Tristate checkbox: all check/all uncheck
            if item.isTristate():
                if state != Qt.PartiallyChecked:
                    self.tree_item_check_all_child(item, check=True if state == Qt.Checked else False)
            # checkbox: all check/all uncheck/partially check
            else:
                self.tree_item_check_child_changed(item)
                self.add_to_run_case_list(item)

    def add_to_run_case_list(self, item):
        full_path_case = self.case_path + self.get_all_parent(item)
        if item.checkState() == Qt.Checked:
            if full_path_case not in self.run_case_list and not os.path.isdir(full_path_case):
                self.run_case_list.append(full_path_case)
                row_case_name = str(full_path_case).split(VAR_SEPARATOR)[-1]
                self.table_model.setItem(self.run_case_list.index(full_path_case), 0, QStandardItem(row_case_name))
        elif full_path_case in self.run_case_list:
            self.table_model.removeRow(self.run_case_list.index(full_path_case))
            self.run_case_list.remove(full_path_case)

    def get_all_parent(self, item):
        parent = item.parent()
        if parent is not None:
            return self.get_all_parent(parent) + VAR_SEPARATOR + item.text()
        else:
            return item.text()

    def get_all_parent_index(self, index):
        parent = index.parent()
        # print(parent.data())
        if parent.data() is not None:
            return self.get_all_parent_index(parent) + VAR_SEPARATOR + index.data()
        else:
            return index.data()

    def tree_item_check_child_changed(self, item):
        sibling_state = self.check_sibling(item)
        parent_item = item.parent()
        if parent_item is None:
            return
        if sibling_state == Qt.PartiallyChecked:
            if parent_item.isCheckable() and parent_item.isTristate():
                parent_item.setCheckState(Qt.PartiallyChecked)
        elif sibling_state == Qt.Checked:
            if parent_item.isCheckable():
                parent_item.setCheckState(Qt.Checked)
        else:
            if parent_item.isCheckable():
                parent_item.setCheckState(Qt.Unchecked)

        self.tree_item_check_child_changed(parent_item)

    @classmethod
    def check_sibling(cls, item):
        parent = item.parent()
        if parent is None:
            return item.checkState()
        brother_count = parent.rowCount()
        checked_count = 0
        unchecked_count = 0
        # there are three types with the child item: all check/all uncheck/partially check
        for i in range(brother_count):
            sibling_item = parent.child(i)
            state = sibling_item.checkState()
            if state == Qt.PartiallyChecked:
                return Qt.PartiallyChecked
            elif state == Qt.Unchecked:
                unchecked_count += 1
            else:
                checked_count += 1
            if checked_count > 0 and unchecked_count > 0:
                return Qt.PartiallyChecked

        if unchecked_count > 0:
            return Qt.Unchecked

        return Qt.Checked

    def tree_item_check_all_child(self, item, check):
        row_count = item.rowCount()
        for i in range(row_count):
            child_items = item.child(i)
            self.tree_item_check_all_child_recursion(child_items, check)
        if item.isCheckable():
            item.setCheckState(Qt.Checked if check else Qt.Unchecked)

    def tree_item_check_all_child_recursion(self, item, check):
        row_count = item.rowCount()
        for i in range(row_count):
            child_items = item.child(i)
            self.tree_item_check_all_child_recursion(child_items, check)
        if item.isCheckable():
            item.setCheckState(Qt.Checked if check else Qt.Unchecked)
            self.add_to_run_case_list(item)

    def create_tree(self, parent, path):
        file_list = os.listdir(path)
        file_list.sort()
        for file in file_list:
            current_path = os.path.join(path, file)
            if os.path.isdir(current_path):
                child_folder = QStandardItem(file)
                child_folder.setCheckable(True)
                child_folder.setTristate(True)
                parent.appendRow(child_folder)
                self.create_tree(child_folder, current_path)
            else:
                if file.__contains__("__init__") or not file.endswith(".py"):
                    continue
                child_file = QStandardItem(file)
                child_file.setCheckable(True)
                parent.appendRow(child_file)


def main():
    app = QApplication(sys.argv)
    window = VatWindow()
    window.show()
    # MainWindow = QMainWindow()
    # ui = Ui_MainWindow()
    # ui.setupUi(MainWindow)
    # MainWindow.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
